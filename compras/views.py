from django.views.generic import DetailView, ListView, TemplateView, RedirectView
from django.shortcuts import redirect, HttpResponse
from django.db.models import Sum
from datetime import datetime
import json

# Model import-->
from compras.models import Compra, DetalleCompra, OfertaCompra
from maestro.models import Proveedor, Almacen
# Model import<--

# Forms import-->
from compras.forms import CompraCreateForm, CompraEditForm, DetalleCompraForm, CompraFiltroForm, ImpuestoForm
from finanzas.forms import PagoCompraForm
# Forms import<--

# Utils import-->
from .utils import fill_data_compra, recalcular_total_compra, \
    cargar_ofertas, create_ofertas, loadtax, cargar_oferta, cancelarcompra
from maestro.utils import empresa_list
# Utils import<--
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Border, Side
from openpyxl import Workbook,load_workbook
# Extra python features-->
# Extra python features<--

# Extra python features-->
from maestro.mixin import BasicEMixin
# Extra python features<--


# Views
class CompraListView(BasicEMixin, ListView):

    template_name = 'compras/compra-list.html'
    model = Compra
    nav_name = 'nav_compra'
    view_name = 'compra'
    action_name = 'leer'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['compra_filtro'] = CompraFiltroForm(self.request.GET, user=self.request.user)
        context['compra_create'] = CompraCreateForm(user=self.request.user)
        return context

    def get_queryset(self):
        query = super().get_queryset()
        alt = 0
        proveedores = self.request.GET.getlist('proveedor')
        estados = self.request.GET.getlist('estado')
        tipo = self.request.GET.getlist('tipo')
        if len(proveedores) > 0:
            query = query.filter(proveedor__in=proveedores)
            alt += 1
        if len(estados) > 0:
            query = query.filter(estado__in=estados)
            alt += 1
        if len(tipo) > 0:
            query = query.filter(tipo__in=tipo)
            alt += 1
        if 'fechahora_creacion1' in self.request.GET and 'fechahora_creacion2' in self.request.GET:
            if self.request.GET['fechahora_creacion1'] != '' and self.request.GET['fechahora_creacion2'] != '':
                alt += 1
                fecha_inicio = datetime.strptime(self.request.GET['fechahora_creacion1'], '%d/%m/%Y %H:%M')
                fecha_fin = datetime.strptime(self.request.GET['fechahora_creacion2'], '%d/%m/%Y %H:%M')
                query = query.filter(fechahora_creacion__gte=fecha_inicio, fechahora_creacion__lte=fecha_fin)
        if 'total_final1' in self.request.GET or 'total_final2' in self.request.GET:
            monto1 = self.request.GET['total_final1']
            monto2 = self.request.GET['total_final2']
            if monto1 == '' and monto2 != '':
                alt += 1
                query = query.filter(total_final__gte=monto1, total_final__lte=monto2)
            if monto1 == '' and monto2 != '':
                alt += 1
                query = query.filter(total_final__lte=monto2)
            elif monto2 == '' and monto1 != '':
                alt += 1
                query = query.filter(total_final__gte=monto1)
        if alt == 0:
            query = Compra.objects.filter(estado__in=[1, 2])
        query = query.filter(proveedor__empresa__in=empresa_list(self.request.user))
        return query


# class OrdenDetailView(BasicEMixin, DetailView):
#
#     template_name = 'compras/orden-detail.html'
#     model = OrdenCompra
#     nav_name = 'nav_compra'
#     view_name = 'orden_compra'
#     action_name = 'leer'
#
#     def get_context_data(self, *, object_list=None, **kwargs):
#         context = super().get_context_data(**kwargs)
#         detalle_orden = DetalleOrdenCompra.objects.filter(ordencompra=self.kwargs['pk'])
#         context['detalle'] = cargar_resultado_oferta(detalle_orden)
#         return context


class CompraCreateView(RedirectView):

    url = '/compras/compra/'
    view_name = 'compra'
    action_name = 'crear'

    def get_redirect_url(self, *args, **kwargs):
        form = CompraCreateForm(self.request.POST, user=self.request.user)
        if form.is_valid():
            try:
                compra = Compra.objects.get(proveedor=form.cleaned_data['proveedor'], estado=1)
            except Compra.DoesNotExist:
                compra = form.save(commit=False)
                compra.asignado = self.request.user
                compra.save()
            url = self.url + str(compra.id) + '/edit'
        else:
            url = '/compras/compra'
        return url


class CompraEditView(BasicEMixin, TemplateView):

    template_name = 'compras/compra-edit.html'
    nav_name = 'nav_compra'
    view_name = 'compra'
    action_name = 'editar'

    def dispatch(self, request, *args, **kwargs):
        compra = Compra.objects.get(pk=self.kwargs['pk'])
        if compra.estado != '1':
            return redirect('/compras/compra/' + str(compra.id))
        else:
            return super().dispatch(request)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        compra = Compra.objects.get(pk=self.kwargs['pk'])
        context['form'] = CompraEditForm(instance=compra)
        context['impuesto_form'] = ImpuestoForm()
        context['model'] = compra
        detalle = DetalleCompra.objects.filter(compra=self.kwargs['pk'], is_oferta=0)
        content_detalle = []
        for idx, d in enumerate(detalle):
            d = cargar_oferta(d)
            d = loadtax(d)
            content_detalle.append([DetalleCompraForm(instance=d, prefix='dc'+str(idx+1),
                                                      proveedor=compra.proveedor_id, has_data=True), d])
        context['detalle'] = content_detalle
        context['clean_form'] = DetalleCompraForm(proveedor=compra.proveedor_id, has_data=False)
        return context

    def post(self, request, *args, **kwargs):
        compra = Compra.objects.get(pk=self.kwargs['pk'])
        form = CompraEditForm(request.POST, instance=compra)
        if form.is_valid():
            compra = form.save()
            DetalleCompra.objects.filter(compra=compra.id, is_oferta=True).delete()
            if request.POST['detallecompra_to_save'] != '':
                total = 0
                for i in request.POST['detallecompra_to_save'].split(','):
                    if 'dc'+i+'-id' in self.request.POST:
                        dc = DetalleCompra.objects.get(pk=self.request.POST['dc'+i+'-id'])
                        dc.impuesto.clear()
                        OfertaCompra.objects.filter(detallecompra=dc.id).delete()
                        dc_form = DetalleCompraForm(request.POST, instance=dc, prefix='dc'+i,
                                                    proveedor=compra.proveedor_id, has_data=True)
                    else:
                        dc_form = DetalleCompraForm(request.POST, prefix='dc'+i,
                                                    proveedor=compra.proveedor_id, has_data=True)
                    if dc_form.is_valid():
                        dc_form = dc_form.save(commit=False)
                        dc_form.compra = compra
                        dc_form.save()
                        create_ofertas(request.POST['dc'+i+'-oferta'], dc_form)
                        fill_data_compra(compra, dc_form, request.POST['dc'+i+'-impuesto'])
                        total += dc_form.total_final
                    else:
                        return HttpResponse(dc_form.errors)
                compra.total_final = total
                if compra.tipo == '1':
                    compra.estado = '3'
                else:
                    compra.estado = '2'
                compra.save()
            if request.POST['detallecompra_to_delete'] != '':
                for j in request.POST['detallecompra_to_delete'].split(','):
                    DetalleCompra.objects.get(pk=j).delete()
        else:
            return HttpResponse(form.errors)
        return redirect('/compras/compra/'+str(compra.id))


class CompraDetailView(BasicEMixin, DetailView):

    template_name = 'compras/compra-detail.html'
    model = Compra
    nav_name = 'nav_compra'
    view_name = 'compra'
    action_name = 'leer'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['detalle'] = cargar_ofertas(DetalleCompra.objects.filter(compra=self.kwargs['pk']))
        context['pago_form'] = PagoCompraForm()
        if 'incidencias' in self.request.GET:
            context['incidencias'] = json.loads(self.request.GET['incidencias'])
        return context


class CompraEntregaView(RedirectView):

    url = '/compras/compra/'
    view_name = 'compra'
    action_name = 'entregar'

    def get_redirect_url(self, *args, **kwargs):
        compra = Compra.objects.get(pk=self.kwargs['compra'])
        compra.is_entregado = True
        compra.save()
        url = self.url + str(compra.id) + ''
        return url


class CompraCancelarView(RedirectView):

    url = '/compras/compra/'
    view_name = 'compra'
    action_name = 'cancelar'

    def get_redirect_url(self, *args, **kwargs):
        compra = Compra.objects.get(pk=self.kwargs['venta'])
        cancelarcompra(compra, self.request.user)
        url = self.url + str(compra.id) + '/edit'
        return url


def reporte_orden(request):
    libro = Workbook()
    libro = load_workbook("./reporteorden.xlsx")
    h = libro.get_sheet_by_name("Hoja1")

    query = Compra.objects.all()
    proveedores = request.POST.getlist('proveedor')
    estados = request.POST.getlist('estado')
    tipo = request.POST.getlist('tipo')
    if len(proveedores) > 0:
        query = query.filter(proveedor__in=proveedores)
    if len(estados) > 0:
        query = query.filter(estado__in=estados)
    if len(tipo) > 0:
        query = query.filter(tipo__in=tipo)
    if 'fechahora_creacion1' in request.POST and 'fechahora_creacion2' in request.POST:
        if request.POST['fechahora_creacion1'] != '' and request.POST['fechahora_creacion2'] != '':
            fecha_inicio = datetime.strptime(request.POST['fechahora_creacion1'], '%d/%m/%Y %H:%M')
            fecha_fin = datetime.strptime(request.POST['fechahora_creacion2'], '%d/%m/%Y %H:%M')
            query = query.filter(fechahora_creacion__gte=fecha_inicio, fechahora_creacion__lte=fecha_fin)
    if 'total_final1' in request.POST or 'total_final2' in request.POST:
        monto1 = request.POST['total_final1']
        monto2 = request.POST['total_final2']
        if monto1 != '' and monto2 != '':
            query = query.filter(total_final__gte=monto1, total_final__lte=monto2)
        if monto1 == '' and monto2 != '':
            query = query.filter(total_final__lte=monto2)
        elif monto2 == '' and monto1 != '':
            query = query.filter(total_final__gte=monto1)
    i = 6
    total = 0
    for o in query:
        i = i + 1
        h.cell(row=i, column=2).value = o.proveedor.descripcion
        h.cell(row=i, column=3).value = o.almacen.descripcion
        h.cell(row=i, column=4).value = o.get_estado_display()
        h.cell(row=i, column=5).value = o.asignado.username
        h.cell(row=i, column=6).value = o.get_tipo_display()
        h.cell(row=i, column=7).value = o.total_final
        total = total + o.total_final

        ### BORDER CADA FILA ###

        h.cell(row=i, column=2).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))
        h.cell(row=i, column=3).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))
        h.cell(row=i, column=4).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))
        h.cell(row=i, column=5).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))
        h.cell(row=i, column=6).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))
        h.cell(row=i, column=7).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))

    i = i + 2
    h.cell(row=i, column=6).value = "TOTAL"
    h.cell(row=i, column=7).value = total

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=compra' + str(o.id) + '.xls'
    libro.save(response)
    return response
