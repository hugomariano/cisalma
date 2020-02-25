from django.views.generic import ListView, DetailView, RedirectView
from django.db.models import Sum
from django.shortcuts import redirect, HttpResponse


# Model import-->
from maestro.models import Almacen, Sucursal, Categoria, Proveedor, Producto
from .models import Stock, Kardex
from compras.models import Compra, DetalleCompra
from ventas.models import Venta, DetalleVenta
from openpyxl import Workbook,load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Border, Side
# Model import<--

# Form import-->
from compras.forms import DetalleCompraRecepcionForm, DetalleCompraNoDeseadoForm, CompraRecepcionForm
from almacen.forms import StockFiltroForm, KardexFiltroForm, RecepcionFiltroForm, KardexReportFiltroForm,\
    StockCambioForm
from ventas.forms import VentaEntregaForm, DetalleVentaEntregaForm
# Form import<--

# Utils import-->
from compras.utils import fill_data_detallecompra
from almacen.utils import loadtax, loadstockdetail
from ventas.utils import fill_data_detalleventa
from maestro.utils import empresa_list
# Utils import<--

# Extra python features-->
from datetime import datetime
# Extra python features<--

# Extra python features-->
from maestro.mixin import BasicEMixin
# Extra python features<--
# Create your views here.


class StockView(BasicEMixin, ListView):

    template_name = 'almacen/stock.html'
    model = Stock
    nav_name = 'nav_stock'
    view_name = 'stock'
    action_name = 'leer'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['stock_filtro'] = StockFiltroForm(self.request.GET, user=self.request.user)
        context['stock_cambio'] = StockCambioForm()
        return context

    def get_queryset(self):
        sucursal = self.request.GET.getlist('sucursal')
        almacen = self.request.GET.getlist('almacen')
        categoria = self.request.GET.getlist('categoria')
        if len(sucursal) > 0:
            query = Stock.objects.filter(almacen__sucursal__in=Sucursal.objects.filter(pk__in=sucursal))
        elif len(almacen) > 0:
            query = Stock.objects.filter(almacen__in=Almacen.objects.filter(pk__in=almacen))
        else:
            query = Stock.objects.all()
        if len(categoria) > 0:
            query = query.filter(producto__categorias__in=categoria)
        query = query.filter(almacen__sucursal__empresa__in=empresa_list(self.request.user))
        query = loadstockdetail(query.values('producto__descripcion', 'producto__id').annotate(Sum('cantidad')))
        return query


class CambiarStockView(RedirectView):

    url = '/almacen/stock/'
    view_name = 'almacen'
    action_name = 'set_stock'

    def get_redirect_url(self, *args, **kwargs):
        form = StockCambioForm(self.request.POST)
        if form.is_valid():
            try:
                producto = Producto.objects.get(descripcion=form.cleaned_data['producto'])
                stock = Stock.objects.get(producto=producto, almacen=1)
                kardex = Kardex.objects.filter(producto=producto, almacen=1)
                diferencia = stock.cantidad - form.cleaned_data['stock']
                stock.cantidad -= diferencia
                stock.save()
                for k in kardex:
                    k.cantidad_saldo -= diferencia
                    k.total_saldo = k.cantidad_saldo*k.precio_unitario_saldo
                    k.save()
            except Producto.DoesNotExist:
                pass
            url = '/almacen/stock/'
        else:
            url = '/almacen/stock/'
        return url


class KardexView(BasicEMixin, ListView):

    template_name = 'almacen/kardex.html'
    model = Kardex
    nav_name = 'nav_kardex'
    view_name = 'stock'
    action_name = 'leer'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['kardex_filtro'] = KardexFiltroForm(self.request.GET, user=self.request.user)
        context['kardex_report'] = KardexReportFiltroForm(self.request.GET, user=self.request.user)
        return context

    def get_queryset(self):
        sucursal = self.request.GET.getlist('sucursales')
        almacen = self.request.GET.getlist('almacenes')
        categoria = self.request.GET.getlist('categorias')
        tipo = self.request.GET.getlist('tipo')
        if len(sucursal) > 0:
            query = Kardex.objects.filter(almacen__sucursal__in=Sucursal.objects.filter(pk__in=sucursal))
        elif len(almacen) > 0:
            query = Kardex.objects.filter(almacen__in=Almacen.objects.filter(pk__in=almacen))
        else:
            query = Kardex.objects.all()
        if len(categoria) > 0:
            query = query.filter(producto__categorias__in=categoria)
        if len(tipo) > 0:
            query = query.filter(tipo_movimiento__in=tipo)
        if 'fecha_inicio' in self.request.GET and 'fecha_fin' in self.request.GET:
            if self.request.GET['fecha_inicio'] != '' and self.request.GET['fecha_fin'] != '':
                fecha_inicio = datetime.strptime(self.request.GET['fecha_inicio'], '%d/%m/%Y %H:%M')
                fecha_fin = datetime.strptime(self.request.GET['fecha_fin'], '%d/%m/%Y %H:%M')
                query = query.filter(fechahora__gte=fecha_inicio, fechahora__lte=fecha_fin)
        query = query.filter(almacen__sucursal__empresa__in=empresa_list(self.request.user))
        query.order_by('fechahora')
        return query


class RecepcionCompraListView(BasicEMixin, ListView):

    template_name = 'almacen/recepcion_compra-list.html'
    model = Compra
    nav_name = 'nav_recepcion_compra'
    view_name = 'recepcion_compra'
    action_name = 'leer'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['recepcion_filtro'] = RecepcionFiltroForm(self.request.GET, user=self.request.user)
        return context

    def get_queryset(self):
        proveedores = self.request.GET.getlist('proveedor')
        query = Compra.objects.filter(tipo='2', estado='2')
        if len(proveedores) > 0:
            query = query.filter(proveedor__in=proveedores)
        query = query.filter(proveedor__empresa__in=empresa_list(self.request.user))
        return query


class RecepcionCompraEditView(BasicEMixin, DetailView):

    template_name = 'almacen/recepcion_compra-edit.html'
    model = Compra
    nav_name = 'nav_recepcion_compra'
    view_name = 'recepcion_compra'
    action_name = 'editar'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        detalle_compra = DetalleCompra.objects.filter(compra=self.kwargs['pk'])
        for d in detalle_compra:
            d = loadtax(d)
        context['detalle'] = detalle_compra
        context['compra_form'] = CompraRecepcionForm(instance=context['object'])
        context['clean_form'] = DetalleCompraNoDeseadoForm(proveedor=context['object'].proveedor_id)
        return context

    def post(self, request, *args, **kwargs):
        compra = Compra.objects.get(pk=self.kwargs['pk'])
        form = CompraRecepcionForm(request.POST, instance=compra)
        if form.is_valid():
            compra = form.save()
            if request.POST['detallecompra_to_save'] != '':
                for i in request.POST['detallecompra_to_save'].split(','):
                    if 'dc'+i+'-id' in self.request.POST:
                        if self.request.POST['dc'+i+'-id'] != '':
                            dc = DetalleCompra.objects.get(pk=self.request.POST['dc'+i+'-id'])
                            dc_form = DetalleCompraRecepcionForm(request.POST, instance=dc, prefix='dc'+i)
                        else:
                            dc_form = DetalleCompraNoDeseadoForm(request.POST, prefix='dc'+i,
                                                                 proveedor=compra.proveedor_id)
                    if dc_form.is_valid():
                        dc_obj = dc_form.save(commit=False)
                        dc_obj.compra = compra
                        fill_data_detallecompra(dc_obj, compra.estado, compra)
            if request.POST['detallecompra_to_delete'] != '':
                for j in request.POST['detallecompra_to_delete'].split(','):
                    detalle_compra = DetalleCompra.objects.get(pk=j)
                    if detalle_compra.is_nodeseado:
                        detalle_compra.delete()
            return redirect('/compras/compra/' + str(compra.id))
        else:
            return HttpResponse(form.errors)


class EntregaVentaListView(BasicEMixin, ListView):

    template_name = 'almacen/entrega_venta-list.html'
    model = Compra
    nav_name = 'nav_entrega_venta'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['entrega_filtro'] = RecepcionFiltroForm(self.request.GET)
        return context

    def get_queryset(self):
        clientes = self.request.GET.getlist('cliente')
        query = Venta.objects.filter(tipo='2', estado='2')
        if len(clientes) > 0:
            query = query.filter(cliente__in=clientes)
        return query


class EntregaVentaEditView(BasicEMixin, DetailView):

    template_name = 'almacen/entrega_venta-edit.html'
    model = Venta
    nav_name = 'nav_entrega_venta'
    view_name = 'entrega_venta'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        detalle_venta = DetalleVenta.objects.filter(venta=self.kwargs['pk'])
        for d in detalle_venta:
            d = loadtax(d)
        context['detalle'] = detalle_venta
        context['venta_form'] = VentaEntregaForm(instance=context['object'])
        return context

    def post(self, request, *args, **kwargs):
        venta = Venta.objects.get(pk=self.kwargs['pk'])
        form = VentaEntregaForm(request.POST, instance=venta)
        if form.is_valid():
            venta = form.save()
            if request.POST['detalleventa_to_save'] != '':
                for i in request.POST['detalleventa_to_save'].split(','):
                    if 'dv'+i+'-id' in self.request.POST:
                        if self.request.POST['dv'+i+'-id'] != '':
                            dv = DetalleVenta.objects.get(pk=self.request.POST['dv'+i+'-id'])
                            dv_form = DetalleVentaEntregaForm(request.POST, instance=dv, prefix='dv'+i)
                    if dv_form.is_valid():
                        dv_obj = dv_form.save(commit=False)
                        dv_obj.venta = venta
                        fill_data_detalleventa(dv_obj, venta.estado, venta)
            return redirect('/ventas/venta/' + str(venta.id))
        else:
            return HttpResponse(form.errors)


def reporte_stock(request):
    libro = Workbook()
    libro = load_workbook("./stock.xlsx")
    h = libro.get_sheet_by_name("Hoja1")

    sucursal = request.POST.getlist('sucursal')
    almacen = request.POST.getlist('almacen')
    categoria = request.POST.getlist('categoria')

    if len(sucursal) > 0:
        query = Stock.objects.filter(almacen__sucursal__in=Sucursal.objects.filter(pk__in=sucursal))
    elif len(almacen) > 0:
        query = Stock.objects.filter(almacen__in=Almacen.objects.filter(pk__in=almacen))
    else:
        query = Stock.objects.all()
    if len(categoria) > 0:
        query = query.filter(producto__categorias__in=categoria)

    query = query.values('producto__descripcion').annotate(Sum('cantidad'))
    i = 6
    for k in query:
        i = i + 1
        h.cell(row=i, column=2).value = k['producto__descripcion']
        h.cell(row=i, column=3).value = k['cantidad__sum']

        h.cell(row=i, column=2).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))
        h.cell(row=i, column=3).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=stock.xls'
        libro.save(response)
    return response


def KardexReportView(request):
    if request.method == 'POST':

        libro = Workbook()
        libro = load_workbook("./kardex.xlsx")
        h = libro.get_sheet_by_name("Hoja1")
        s = Sucursal.objects.get(id=request.POST['sucursal'])
        try:
            p = Producto.objects.get(id=request.POST['productos'], empresa_id=s.empresa_id)
            fecha_inicio = datetime.strptime(request.POST['date_inicio'], '%d/%m/%Y %H:%M')
            fecha_fin = datetime.strptime(request.POST['date_fin'], '%d/%m/%Y %H:%M')
            # k=Kardex.objects.filter(
            #   fechahora__gte=fecha_inicio, fechahora__lte=fecha_fin,
            #   producto_id=p.id)
            k = Kardex.objects.filter(producto_id=p.id)

            i = 13
            sum = 0
            h.cell(row=2, column=4).value = fecha_inicio
            h.cell(row=3, column=4).value = p.empresa.ruc
            h.cell(row=4, column=4).value = p.empresa.descripcion
            h.cell(row=5, column=4).value = s.descripcion
            h.cell(row=6, column=4).value = p.id
            h.cell(row=8, column=4).value = p.descripcion
            for o in k:
                i = i + 1
                h.cell(row=i, column=2).value = o.fechahora
                h.cell(row=i, column=3).value = o.get_tipo_movimiento_display()
                h.cell(row=i, column=4).value = o.serie_comprobante
                h.cell(row=i, column=5).value = o.numero_comprobante
                h.cell(row=i, column=6).value = o.get_tipo_detalle_display()
                h.cell(row=i, column=7).value = o.cantidad_entrada
                h.cell(row=i, column=8).value = o.precio_unitario_entrada
                h.cell(row=i, column=9).value = o.total_entrada
                h.cell(row=i, column=10).value = o.cantidad_salida
                h.cell(row=i, column=11).value = o.precio_unitario_salida
                h.cell(row=i, column=12).value = o.total_salida
                h.cell(row=i, column=13).value = o.cantidad_saldo
                h.cell(row=i, column=14).value = o.precio_unitario_saldo
                h.cell(row=i, column=15).value = o.total_saldo

                ### BORDER CADA FILA ###
                h.cell(row=i, column=2).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                        right=Side(border_style='thin', color='FF000000'),
                                                        bottom=Side(border_style='thin', color='FF000000'),
                                                        left=Side(border_style='thin', color='FF000000'))
                h.cell(row=i, column=3).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                        right=Side(border_style='thin', color='FF000000'),
                                                        bottom=Side(border_style='thin', color='FF000000'),
                                                        left=Side(border_style='thin', color='FF000000'))
                h.cell(row=i, column=4).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                        right=Side(border_style='thin', color='FF000000'),
                                                        bottom=Side(border_style='thin', color='FF000000'),
                                                        left=Side(border_style='thin', color='FF000000'))
                h.cell(row=i, column=5).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                        right=Side(border_style='thin', color='FF000000'),
                                                        bottom=Side(border_style='thin', color='FF000000'),
                                                        left=Side(border_style='thin', color='FF000000'))
                h.cell(row=i, column=6).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                        right=Side(border_style='thin', color='FF000000'),
                                                        bottom=Side(border_style='thin', color='FF000000'),
                                                        left=Side(border_style='thin', color='FF000000'))
                h.cell(row=i, column=7).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                        right=Side(border_style='thin', color='FF000000'),
                                                        bottom=Side(border_style='thin', color='FF000000'),
                                                        left=Side(border_style='thin', color='FF000000'))
                h.cell(row=i, column=8).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                        right=Side(border_style='thin', color='FF000000'),
                                                        bottom=Side(border_style='thin', color='FF000000'),
                                                        left=Side(border_style='thin', color='FF000000'))
                h.cell(row=i, column=9).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                        right=Side(border_style='thin', color='FF000000'),
                                                        bottom=Side(border_style='thin', color='FF000000'),
                                                        left=Side(border_style='thin', color='FF000000'))
                h.cell(row=i, column=10).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                         right=Side(border_style='thin', color='FF000000'),
                                                         bottom=Side(border_style='thin', color='FF000000'),
                                                         left=Side(border_style='thin', color='FF000000'))
                h.cell(row=i, column=11).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                         right=Side(border_style='thin', color='FF000000'),
                                                         bottom=Side(border_style='thin', color='FF000000'),
                                                         left=Side(border_style='thin', color='FF000000'))
                h.cell(row=i, column=12).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                         right=Side(border_style='thin', color='FF000000'),
                                                         bottom=Side(border_style='thin', color='FF000000'),
                                                         left=Side(border_style='thin', color='FF000000'))
                h.cell(row=i, column=13).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                         right=Side(border_style='thin', color='FF000000'),
                                                         bottom=Side(border_style='thin', color='FF000000'),
                                                         left=Side(border_style='thin', color='FF000000'))
                h.cell(row=i, column=14).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                         right=Side(border_style='thin', color='FF000000'),
                                                         bottom=Side(border_style='thin', color='FF000000'),
                                                         left=Side(border_style='thin', color='FF000000'))
                h.cell(row=i, column=15).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                         right=Side(border_style='thin', color='FF000000'),
                                                         bottom=Side(border_style='thin', color='FF000000'),
                                                         left=Side(border_style='thin', color='FF000000'))

            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=kardex.xls'
            libro.save(response)
            return response

        except Producto.DoesNotExist:
            return HttpResponse('NO EXISTE REGISTROS')

