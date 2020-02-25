from django.views.generic import DetailView, ListView, TemplateView, RedirectView, UpdateView
from django.shortcuts import redirect, HttpResponse
import json
# Mixin Import-->
from maestro.mixin import BasicEMixin
# Mixin Import<--

# Extra python features-->
from datetime import datetime
# Extra python features<--

from .utils import fill_data_venta, load_tax, create_venta_txt, cancelarventa
from maestro.utils import empresa_list

# Model import-->
from ventas.models import OfertaVenta, Venta, DetalleVenta
# Model import<--

# Forms import-->
from .forms import OfertaVentaForm, VentaFiltroForm, VentaCreateForm, VentaEditForm, DetalleVentaForm, ImpuestoForm,\
    VentaDescuentoAdicionalForm
from finanzas.forms import PagoVentaForm
# Forms import<--
from openpyxl.styles import Border, Side
from openpyxl import Workbook, load_workbook


# Create your views here.
class OfertaListView(BasicEMixin, ListView):

    template_name = 'ventas/oferta-list.html'
    model = OfertaVenta
    nav_name = 'nav_oferta'
    view_name = 'oferta_venta'
    action_name = 'leer'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

    def get_queryset(self):
        categorias = self.request.GET.getlist('categoria')
        if len(categorias) > 0:
            query = OfertaVenta.objects.filter(categorias__in=categorias)
        else:
            query = OfertaVenta.objects.all()
        query = query.filter(sucursal__empresa__in=empresa_list(self.request.user))
        return query


class OfertaDetailView(BasicEMixin, DetailView):

    template_name = 'ventas/oferta-detail.html'
    model = OfertaVenta
    nav_name = 'nav_oferta'
    view_name = 'oferta_venta'
    action_name = 'leer'


class OfertaEditView(BasicEMixin, TemplateView):

    template_name = 'ventas/oferta-edit.html'
    nav_name = 'nav_oferta'
    view_name = 'oferta_venta'
    action_name = 'editar'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.kwargs['pk'] == 0:
                context['object'] = OfertaVentaForm(user=self.request.user)
        else:
            producto = OfertaVenta.objects.get(pk=self.kwargs['pk'])
            context['object'] = OfertaVentaForm(instance=producto, user=self.request.user)
        return context

    def post(self, request, *args, **kwargs):
        if self.kwargs['pk'] == 0:
            form = OfertaVentaForm(request.POST, user=self.request.user)
        else:
            producto = OfertaVenta.objects.get(pk=self.kwargs['pk'])
            form = OfertaVentaForm(request.POST, instance=producto, user=self.request.user)
        if form.is_valid():
            oferta = form.save(commit=False)
            if oferta.stock_faltante is None:
                oferta.stock_faltante = oferta.stock_limite
            oferta.save()
        else:
            return HttpResponse(form.errors)
        return redirect('/ventas/oferta/' + str(oferta.id))


class VentaListView(BasicEMixin, ListView):

    template_name = 'ventas/venta-list.html'
    model = Venta
    nav_name = 'nav_venta'
    view_name = 'venta'
    action_name = 'leer'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['venta_filtro'] = VentaFiltroForm(self.request.GET, user=self.request.user)
        context['venta_create'] = VentaCreateForm(user=self.request.user)
        return context

    def get_queryset(self):
        alt = 0
        query = super().get_queryset()
        cliente = self.request.GET.getlist('cliente')
        sucursal = self.request.GET.getlist('sucursal')
        estado = self.request.GET.getlist('estado')
        tipo_pago = self.request.GET.getlist('tipo_pago')
        estado_pago = self.request.GET.getlist('estado_pago')
        if len(cliente) > 0:
            alt += 1
            query = query.filter(cliente__in=cliente)
        if len(sucursal) > 0:
            alt += 1
            query = query.filter(sucursal__in=sucursal)
        if len(estado) > 0:
            alt += 1
            query = query.filter(estado__in=estado)
        if len(tipo_pago) > 0:
            alt += 1
            query = query.filter(tipo_pago__in=tipo_pago)
        if len(estado_pago) > 0:
            alt += 1
            query = query.filter(estado_pago__in=estado_pago)
        if 'fechahora_creacion1' in self.request.GET and 'fechahora_creacion2' in self.request.GET:
            if self.request.GET['fechahora_creacion1'] != '' and self.request.GET['fechahora_creacion2'] != '':
                alt += 1
                fecha_inicio = datetime.strptime(self.request.GET['fechahora_creacion1'], '%d/%m/%Y %H:%M')
                fecha_fin = datetime.strptime(self.request.GET['fechahora_creacion2'], '%d/%m/%Y %H:%M')
                query = query.filter(fechahora_creacion__gte=fecha_inicio, fechahora_creacion__lte=fecha_fin)
        if 'total1' in self.request.GET or 'total2' in self.request.GET:
            total1 = self.request.GET['total1']
            total2 = self.request.GET['total2']
            if total1 != '' and total2 != '':
                alt += 1
                query = query.filter(total__gte=total1, total__lte=total2)
            elif total1 == '' and total2 != '':
                alt += 1
                query = query.filter(total__lte=total2)
            elif total2 == '' and total1 != '':
                alt += 1
                query = query.filter(total__gte=total1)
        if alt == 0:
            query = Venta.objects.filter(estado__in=[1, 2])
        query = query.filter(sucursal__empresa__in=empresa_list(self.request.user))
        return query


class VentaCreateView(BasicEMixin, RedirectView):

    url = '/ventas/venta/'
    view_name = 'venta'
    action_name = 'crear'

    def get_redirect_url(self, *args, **kwargs):
        form = VentaCreateForm(self.request.POST, user=self.request.user)
        if form.is_valid():
            if form.cleaned_data['cliente'] is not None:
                try:
                    venta = Venta.objects.get(cliente=form.cleaned_data['cliente'], estado=1)
                except Venta.DoesNotExist:
                    venta = form.save(commit=False)
                    venta.asignado = self.request.user
                    venta.save()
            else:
                venta = form.save(commit=False)
                venta.asignado = self.request.user
                venta.save()
            url = self.url + str(venta.id) + '/edit'
        else:
            url = '/ventas/venta'
        return url


class VentaDuplicarView(BasicEMixin, RedirectView):

    url = '/ventas/venta/'
    view_name = 'venta'
    action_name = 'crear'

    def get_redirect_url(self, *args, **kwargs):
        venta = Venta.objects.get(pk=self.kwargs['venta'])
        detalleventa = DetalleVenta.objects.filter(venta=venta.id, is_oferta=False)
        venta.estado = '1'
        venta.pk = None
        venta.save()
        for d in detalleventa:
            d.pk = None
            d.venta = venta
            d.save()
        url = self.url + str(venta.id) + '/edit'
        return url


class VentaDescuentoAdicionalView(BasicEMixin, UpdateView):

    form_class = VentaDescuentoAdicionalForm
    model = DetalleVenta
    success_url = '/ventas/venta/'
    view_name = 'venta'
    action_name = 'descuento_adicional'

    def form_valid(self, form):
        detalleventa = form.save(commit=False)
        if not detalleventa.venta.is_pagado:
            detalleventa.total_final -= detalleventa.descuento_adicional
            detalleventa.save()
            venta = detalleventa.venta
            venta.descuento += detalleventa.descuento_adicional
            venta.total_final -= detalleventa.descuento_adicional
            venta.save()
        return super().form_valid(form)

    def get_success_url(self):
        url = self.success_url + str(self.object.id)
        return url


class VentaEditView(BasicEMixin, TemplateView):

    template_name = 'ventas/venta-edit.html'
    nav_name = 'nav_venta'
    view_name = 'venta'
    action_name = 'editar'

    def dispatch(self, request, *args, **kwargs):
        venta = Venta.objects.get(pk=self.kwargs['pk'])
        if venta.estado != '1':
            return redirect('/ventas/venta/' + str(venta.id))
        else:
            return super().dispatch(request)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        venta = Venta.objects.get(pk=self.kwargs['pk'])
        context['form'] = VentaEditForm(instance=venta, user=self.request.user)
        context['impuesto_form'] = ImpuestoForm()
        context['model'] = venta
        detalle = DetalleVenta.objects.filter(venta=self.kwargs['pk'], is_oferta=False)
        content_detalle = []
        for idx, d in enumerate(detalle):
            d = load_tax(d)
            content_detalle.append([DetalleVentaForm(instance=d, prefix='dv'+str(idx+1),
                                                     sucursal=venta.sucursal_id, has_data=True), d])
        context['detalle'] = content_detalle
        context['clean_form'] = DetalleVentaForm(sucursal=venta.sucursal_id, has_data=False)
        return context

    def post(self, request, *args, **kwargs):
        venta = Venta.objects.get(pk=self.kwargs['pk'])
        form = VentaEditForm(request.POST, instance=venta, user=self.request.user)
        incidencias = []
        if form.is_valid():
            venta = form.save()
            DetalleVenta.objects.filter(venta=venta.id, is_oferta=True).delete()
            if request.POST['detalleventa_to_save'] != '':
                total = 0
                sub_total = 0
                total_final = 0
                descuento = 0
                for i in request.POST['detalleventa_to_save'].split(','):
                    if 'dv'+i+'-id' in self.request.POST:
                        dv = DetalleVenta.objects.get(pk=self.request.POST['dv'+i+'-id'])
                        dv.impuesto.clear()
                        dv_form = DetalleVentaForm(request.POST, instance=dv, prefix='dv'+i,
                                                   sucursal=venta.sucursal_id, has_data=True)
                    else:
                        dv_form = DetalleVentaForm(request.POST, prefix='dv'+i,
                                                   sucursal=venta.sucursal_id, has_data=True)
                    if dv_form.is_valid():
                        dv_form = dv_form.save(commit=False)
                        dv_form.venta = venta
                        incidencia = fill_data_venta(venta, dv_form, request.POST['dv'+i+'-impuesto_inp'])
                        if len(incidencia) > 0:
                            incidencias.append(incidencia)
                        total_final += dv_form.total_final
                        total += dv_form.total
                        sub_total += dv_form.sub_total
                        descuento += dv_form.descuento
                    else:
                        return HttpResponse(dv_form.errors)
                venta.total_final = total_final
                venta.total = total
                venta.sub_total = sub_total
                venta.descuento = descuento
                venta.save()
            if venta.tipo == '1':
                venta.estado = '3'
            else:
                venta.estado = '2'
            venta.save()
            if request.POST['detalleventa_to_delete'] != '':
                for j in request.POST['detalleventa_to_delete'].split(','):
                    DetalleVenta.objects.get(pk=j).delete()
        else:
            return HttpResponse(form.errors)
        return redirect('/ventas/venta/'+str(venta.id)+'/?incidencias='+json.dumps(incidencias))


class VentaDetailView(BasicEMixin, DetailView):

    template_name = 'ventas/venta-detail.html'
    model = Venta
    nav_name = 'nav_venta'
    view_name = 'venta'
    action_name = 'leer'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        create_venta_txt(self.kwargs['pk'])
        context['detalle'] = DetalleVenta.objects.filter(venta=self.kwargs['pk'])
        context['pago_form'] = PagoVentaForm()
        context['descuento_form'] = VentaDescuentoAdicionalForm()
        if 'incidencias' in self.request.GET:
            context['incidencias'] = json.loads(self.request.GET['incidencias'])
        return context


class VentaEntregaView(RedirectView):

    url = '/ventas/venta/'
    view_name = 'venta'
    action_name = 'entregar'

    def get_redirect_url(self, *args, **kwargs):
        venta = Venta.objects.get(pk=self.kwargs['venta'])
        venta.is_entregado = True
        venta.save()
        url = self.url + str(venta.id) + ''
        return url


class VentaCancelarView(RedirectView):

    url = '/ventas/venta/'
    view_name = 'venta'
    action_name = 'cancelar'

    def get_redirect_url(self, *args, **kwargs):
        venta = Venta.objects.get(pk=self.kwargs['venta'])
        resp = cancelarventa(venta, self.request.user)
        url = self.url + str(venta.id) + '/'+resp
        return url


class VentaFindView(RedirectView):

    url = '/ventas/venta'
    view_name = 'ventas'
    action_name = 'leer'

    def get_redirect_url(self, *args, **kwargs):
        try:
            venta = Venta.objects.get(pk=self.request.GET['venta'])
            url = self.url + '/' + self.request.GET['venta']
        except Venta.DoesNotExist:
            url = self.url
        return url


def ReporteVentas(request, id):
    libro = Workbook()
    libro = load_workbook("./detalleventa.xlsx")
    h = libro.get_sheet_by_name("Hoja1")
    venta = Venta.objects.get(id=id)
    detalleventa = DetalleVenta.objects.filter(venta_id=venta.id)

    i = 7
    sum = 0
    fecha = str(venta.fechahora_creacion)
    new_fecha = fecha[:19]
    h.cell(row=3, column=3).value = str(venta.sucursal.descripcion)
    h.cell(row=3, column=5).value = str(venta.get_estado_display())
    h.cell(row=3, column=8).value = new_fecha
    h.cell(row=5, column=3).value = venta.total
    h.cell(row=5, column=5).value = venta.asignado.username
    h.cell(row=5, column=8).value = str(venta.cliente)

    for dv in detalleventa:
        i = i + 1
        h.cell(row=i, column=2).value = dv.producto.descripcion
        h.cell(row=i, column=2).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))
        h.cell(row=i, column=3).value = str(dv.presentacionxproducto)
        h.cell(row=i, column=3).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))
        h.cell(row=i, column=4).value = str(dv.cantidad_presentacion_pedido) + ' ' + ' ≅ ' + ' ' + str(
            dv.cantidad_unidad_pedido)
        h.cell(row=i, column=4).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))
        h.cell(row=i, column=5).value = dv.precio
        h.cell(row=i, column=5).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))
        h.cell(row=i, column=6).value = dv.total
        h.cell(row=i, column=6).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))
        h.cell(row=i, column=7).value = dv.descuento
        h.cell(row=i, column=7).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))
        h.cell(row=i, column=8).value = dv.total_final
        h.cell(row=i, column=8).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))
        h.cell(row=i, column=9).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=detalleventa' + str(venta.id) + '.xls'
    libro.save(response)
    return response
