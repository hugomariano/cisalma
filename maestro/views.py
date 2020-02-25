from django.views.generic import DetailView, ListView, TemplateView, RedirectView
from django.shortcuts import redirect, HttpResponse
from django.db.models import Sum
from django.contrib.auth.hashers import make_password
# Model import-->
from maestro.models import Empresa, Sucursal, Almacen, Categoria,\
    Presentacion, Producto, PresentacionxProducto, Proveedor, CatalogoxProveedor, Caja, AsignacionGrupo, Grupo
from ventas.models import OfertaVenta
from django.contrib.auth.models import User
# Model import<--

# Forms import-->
from .forms import SucursalForm, AlmacenForm, CategoriaForm, PresentacionForm,\
    ProductoForm, ProductoCategoriaForm, ProductoPresentacionForm, ProveedorForm,\
    CatalogoProveedorForm, ProductoFiltroForm, CatalogoFiltroForm, CatalogoProveedorFiltroForm, CajaForm, UsuarioForm, \
    EmpresaForm
# Forms import<--

# Utils import-->
from .utils import format_categories, empresa_list
# Utils import<--

# Extra python features-->
# Extra python features<--

# Extra python features-->
from .mixin import BasicEMixin
# Extra python features<--
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Border, Side
from openpyxl import Workbook,load_workbook


# Views
class EmpresaListView(BasicEMixin, ListView):

    template_name = 'maestro/empresa-list.html'
    model = Empresa
    nav_name = 'nav_empresa'
    view_name = 'empresa'
    action_name = 'leer'


class EmpresaDetailView(BasicEMixin, DetailView):

    template_name = 'maestro/empresa-detail.html'
    model = Empresa
    nav_name = 'nav_empresa'
    view_name = 'empresa'
    action_name = 'leer'


class EmpresaEditView(BasicEMixin, TemplateView):

    template_name = 'maestro/empresa-edit.html'
    nav_name = 'nav_empresa'
    view_name = 'empresa'
    action_name = 'editar'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.kwargs['pk'] == 0:
            context['object'] = EmpresaForm()
        else:
            empresa = Empresa.objects.get(pk=self.kwargs['pk'])
            context['object'] = EmpresaForm(instance=empresa)
        return context

    def post(self, request, *args, **kwargs):
        if self.kwargs['pk'] == 0:
            form = EmpresaForm(request.POST)
        else:
            empresa = Empresa.objects.get(pk=self.kwargs['pk'])
            form = EmpresaForm(request.POST, instance=empresa)
        if form.is_valid():
            form.save()
        else:
            return HttpResponse(form.errors)
        return redirect('/maestro/empresa')


class SucursalListView(BasicEMixin, ListView):

    template_name = 'maestro/sucursal-list.html'
    model = Sucursal
    nav_name = 'nav_sucursal'
    view_name = 'sucursal'
    action_name = 'leer'


class SucursalDetailView(BasicEMixin, DetailView):

    template_name = 'maestro/sucursal-detail.html'
    model = Sucursal
    nav_name = 'nav_sucursal'
    view_name = 'sucursal'
    action_name = 'leer'


class SucursalEditView(BasicEMixin, TemplateView):

    template_name = 'maestro/sucursal-edit.html'
    nav_name = 'nav_sucursal'
    view_name = 'sucursal'
    action_name = 'editar'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.kwargs['pk'] == 0:
            context['object'] = SucursalForm()
        else:
            sucursal = Sucursal.objects.get(pk=self.kwargs['pk'])
            context['object'] = SucursalForm(instance=sucursal)
        return context

    def post(self, request, *args, **kwargs):
        if self.kwargs['pk'] == 0:
            form = SucursalForm(request.POST)
        else:
            sucursal = Sucursal.objects.get(pk=self.kwargs['pk'])
            form = SucursalForm(request.POST, instance=sucursal)
        if form.is_valid():
            form.save()
        else:
            return HttpResponse(form.errors)
        return redirect('/maestro/sucursal')


class AlmacenListView(BasicEMixin, ListView):

    template_name = 'maestro/almacen-list.html'
    model = Almacen
    nav_name = 'nav_almacen'
    view_name = 'almacen'
    action_name = 'leer'


class AlmacenDetailView(BasicEMixin, DetailView):

    template_name = 'maestro/almacen-detail.html'
    model = Almacen
    nav_name = 'nav_almacen'
    view_name = 'almacen'
    action_name = 'leer'


class AlmacenEditView(BasicEMixin, TemplateView):

    template_name = 'maestro/almacen-edit.html'
    nav_name = 'nav_almacen'
    view_name = 'almacen'
    action_name = 'editar'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.kwargs['pk'] == 0:
            context['object'] = AlmacenForm()
        else:
            almacen = Almacen.objects.get(pk=self.kwargs['pk'])
            context['object'] = AlmacenForm(instance=almacen)
        return context

    def post(self, request, *args, **kwargs):
        if self.kwargs['pk'] == 0:
            form = AlmacenForm(request.POST)
        else:
            almacen = Almacen.objects.get(pk=self.kwargs['pk'])
            form = AlmacenForm(request.POST, instance=almacen)
        if form.is_valid():
            form.save()
        else:
            return HttpResponse(form.errors)
        return redirect('/maestro/almacen')


class CajaListView(BasicEMixin, ListView):

    template_name = 'maestro/caja-list.html'
    model = Caja
    nav_name = 'nav_caja'
    view_name = 'caja'
    action_name = 'leer'


class CajaDetailView(BasicEMixin, DetailView):

    template_name = 'maestro/caja-detail.html'
    model = Caja
    nav_name = 'nav_caja'
    view_name = 'caja'
    action_name = 'leer'


class CajaEditView(BasicEMixin, TemplateView):

    template_name = 'maestro/caja-edit.html'
    nav_name = 'nav_caja'
    view_name = 'caja'
    action_name = 'editar'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.kwargs['pk'] == 0:
            context['object'] = CajaForm()
        else:
            caja = Caja.objects.get(pk=self.kwargs['pk'])
            context['object'] = CajaForm(instance=caja)
        return context

    def post(self, request, *args, **kwargs):
        if self.kwargs['pk'] == 0:
            form = CajaForm(request.POST)
        else:
            caja = Caja.objects.get(pk=self.kwargs['pk'])
            form = CajaForm(request.POST, instance=caja)
        if form.is_valid():
            form.save()
        else:
            return HttpResponse(form.errors)
        return redirect('/maestro/caja')


class CategoriaListView(BasicEMixin, ListView):

    template_name = 'maestro/categoria-list.html'
    model = Categoria
    nav_name = 'nav_categoria'
    nav_main = 'nav_main_producto'
    view_name = 'categoria'
    action_name = 'leer'


class CategoriaDetailView(BasicEMixin, DetailView):

    template_name = 'maestro/categoria-detail.html'
    model = Categoria
    nav_name = 'nav_categoria'
    nav_main = 'nav_main_producto'
    view_name = 'categoria'
    action_name = 'leer'


class CategoriaEditView(BasicEMixin, TemplateView):

    template_name = 'maestro/categoria-edit.html'
    nav_name = 'nav_categoria'
    nav_main = 'nav_main_producto'
    view_name = 'categoria'
    action_name = 'editar'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.kwargs['pk'] == 0:
            context['object'] = CategoriaForm()
        else:
            categoria = Categoria.objects.get(pk=self.kwargs['pk'])
            context['object'] = CategoriaForm(instance=categoria)
        return context

    def post(self, request, *args, **kwargs):
        if self.kwargs['pk'] == 0:
            form = CategoriaForm(request.POST)
        else:
            categoria = Categoria.objects.get(pk=self.kwargs['pk'])
            form = CategoriaForm(request.POST, instance=categoria)
        if form.is_valid():
            categoria = form.save(commit=False)
            if form.cleaned_data['padre'] is None:
                categoria.nivel = 1
                categoria.save()
                categoria.padre_total = categoria
                categoria.save()
            else:
                padre = Categoria.objects.get(pk=form.cleaned_data['padre'].pk)
                categoria.nivel = padre.nivel + 1
                categoria.padre_total = padre.padre_total
                categoria.save()
        else:
            return HttpResponse(form.errors)
        return redirect('/maestro/categoria')


class PresentacionListView(BasicEMixin, ListView):

    template_name = 'maestro/presentacion-list.html'
    model = Presentacion
    nav_name = 'nav_presentacion'
    nav_main = 'nav_main_producto'
    view_name = 'presentacion'
    action_name = 'leer'


class PresentacionDetailView(BasicEMixin, DetailView):

    template_name = 'maestro/presentacion-detail.html'
    model = Presentacion
    nav_name = 'nav_presentacion'
    nav_main = 'nav_main_producto'
    view_name = 'presentacion'
    action_name = 'leer'


class PresentacionEditView(BasicEMixin, TemplateView):

    template_name = 'maestro/presentacion-edit.html'
    nav_name = 'nav_presentacion'
    nav_main = 'nav_main_producto'
    view_name = 'presentacion'
    action_name = 'editar'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.kwargs['pk'] == 0:
            context['object'] = PresentacionForm()
        else:
            presentacion = Presentacion.objects.get(pk=self.kwargs['pk'])
            context['object'] = PresentacionForm(instance=presentacion)
        return context

    def post(self, request, *args, **kwargs):
        if self.kwargs['pk'] == 0:
            form = PresentacionForm(request.POST)
        else:
            presentacion = Presentacion.objects.get(pk=self.kwargs['pk'])
            form = PresentacionForm(request.POST, instance=presentacion)
        if form.is_valid():
            form.save()
        else:
            return HttpResponse(form.errors)
        return redirect('/maestro/presentacion')


class ProductoListView(BasicEMixin, ListView):

    template_name = 'maestro/producto-list.html'
    model = Producto
    nav_name = 'nav_producto'
    nav_main = 'nav_main_producto'
    view_name = 'producto'
    action_name = 'leer'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['producto_filtro'] = ProductoFiltroForm(self.request.GET)
        return context

    def get_queryset(self):
        categorias = self.request.GET.getlist('categoria')
        if len(categorias) > 0:
            query = Producto.objects.filter(categorias__in=categorias)
        else:
            query = Producto.objects.all()
        return query


class ProductoDetailView(BasicEMixin, DetailView):

    template_name = 'maestro/producto-detail.html'
    model = Producto
    nav_name = 'nav_producto'
    nav_main = 'nav_main_producto'
    view_name = 'producto'
    action_name = 'leer'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categorias'] = format_categories(Categoria.objects.filter(producto=self.kwargs['pk'])
                                                  .order_by('padre_total', 'nivel'))
        context['presentaciones'] = PresentacionxProducto.objects.filter(producto=self.kwargs['pk'])
        context['catalogo'] = Sucursal.objects.filter(producto=self.kwargs['pk'])
        return context


class ProductoPrecioView(BasicEMixin, TemplateView):

    template_name = 'maestro/producto-precio.html'
    nav_name = 'nav_producto'
    nav_main = 'nav_main_producto'
    view_name = 'producto'
    action_name = 'set_precio'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['object'] = Producto.objects.get(pk=self.kwargs['pk'])
        context['presentaciones'] = PresentacionxProducto.objects.filter(producto=self.kwargs['pk'])
        return context

    def post(self, request, *args, **kwargs):
        producto = Producto.objects.get(pk=self.kwargs['pk'])
        OfertaVenta.objects.filter(producto_oferta=producto).delete()
        presentacionxproducto = PresentacionxProducto.objects.filter(producto=producto)
        sucursal = Sucursal.objects.get(pk=1)
        for p in presentacionxproducto:
            if request.POST['precio_compra-'+str(p.id)] != '':
                precio_compra = float(request.POST['precio_compra-'+str(p.id)]) / p.cantidad
                if producto.precio_compra < precio_compra:
                    producto.precio_compra = precio_compra
            if request.POST['precio_venta-'+str(p.id)] != '':
                precio_venta = round(float(request.POST['precio_venta-'+str(p.id)]) / p.cantidad, 1)
                if producto.precio_venta <= precio_venta:
                    producto.precio_venta = precio_venta
                ofertaventa = OfertaVenta.objects.filter(producto_oferta=producto, is_active=True, estado=True,
                                                         cantidad_unidad_oferta__lte=p.cantidad)
                descuento = 0
                for ofv in ofertaventa:
                    descuento += (p.cantidad/ofv.presentacion_oferta.cantidad) * float(ofv.retorno)
                if (p.cantidad * producto.precio_venta) - descuento > float(request.POST['precio_venta-'+str(p.id)]):
                    OfertaVenta(sucursal=sucursal, tipo='2', tipo_duracion='2', producto_oferta=producto,
                                presentacion_oferta=p, cantidad_oferta=1,
                                cantidad_unidad_oferta=p.cantidad,
                                retorno=(float(producto.precio_venta * p.cantidad)-float(request.POST['precio_venta-'+str(p.id)]))-descuento).save()
        producto.utilidad_monetaria = float(producto.precio_venta) - float(producto.precio_compra)
        producto.save()
        next_prod = Producto.objects.filter(precio_venta=0).order_by('id')[:1]
        if len(next_prod) == 0:
            return redirect('/maestro/producto/'+str(producto.id))
        else:
            return redirect('/maestro/producto/'+str(producto.id)+'/precio')


class ProductoEditView(BasicEMixin, TemplateView):

    template_name = 'maestro/producto-edit.html'
    nav_name = 'nav_producto'
    nav_main = 'nav_main_producto'
    view_name = 'producto'
    action_name = 'editar'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.kwargs['pk'] == 0:
                context['object'] = ProductoForm
        else:
            producto = Producto.objects.get(pk=self.kwargs['pk'])
            context['object'] = ProductoForm(instance=producto)
        return context

    def post(self, request, *args, **kwargs):
        if self.kwargs['pk'] == 0:
            form = ProductoForm(request.POST)
        else:
            producto = Producto.objects.get(pk=self.kwargs['pk'])
            form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
        else:
            return HttpResponse(form.errors)
        return redirect('/maestro/producto')


class ProductoCategoriaView(BasicEMixin, TemplateView):

    template_name = 'maestro/producto-categoria.html'
    nav_name = 'nav_producto'
    nav_main = 'nav_main_producto'
    view_name = 'producto'
    action_name = 'set_categoria'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pk = self.kwargs['pk']
        producto = Producto.objects.get(pk=pk)
        context['object'] = ProductoCategoriaForm(instance=producto)
        context['categorias'] = format_categories(Categoria.objects.filter(producto=self.kwargs['pk'])
                                                  .order_by('padre_total', 'nivel'))
        return context

    def post(self, request, *args, **kwargs):
        pk = self.kwargs['pk']
        producto = Producto.objects.get(pk=pk)
        form = ProductoCategoriaForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
        else:
            return HttpResponse(form.errors)
        return redirect('/maestro/producto/'+str(pk))


class ProductoPresentacionView(BasicEMixin, TemplateView):

    template_name = 'maestro/producto-presentacion.html'
    nav_name = 'nav_producto'
    nav_main = 'nav_main_producto'
    view_name = 'producto'
    action_name = 'set_presentacion'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pk = self.kwargs['pk']
        context['presentaciones'] = Presentacion.objects.all()
        context['own_presentaciones'] = PresentacionxProducto.objects.filter(producto=pk)
        return context

    def post(self, request, *args, **kwargs):
        pk = self.kwargs['pk']
        producto = Producto.objects.get(pk=pk)
        if request.POST['presentacion_to_save'] != "":
            presentacion_toadd = request.POST['presentacion_to_save'].split(',')
            for p in presentacion_toadd:
                if request.POST['p'+p+'-id'] != '':
                    presentacionxproducto = PresentacionxProducto.objects.get(pk=request.POST['p'+p+'-id'])
                    form = ProductoPresentacionForm(request.POST, instance=presentacionxproducto, prefix='p'+p)
                else:
                    form = ProductoPresentacionForm(request.POST, prefix='p'+p)
                if form.is_valid():
                    presentacionxproducto = form.save(commit=False)
                    presentacionxproducto.producto = producto
                    presentacionxproducto.save()
                else:
                    return HttpResponse(form.errors)
        if request.POST['presentacion_to_delete'] != "":
            presentacion_todelete = request.POST['presentacion_to_delete'].split(',')
            for p in presentacion_todelete:
                PresentacionxProducto.objects.get(pk=p).delete()
        return redirect('/maestro/producto/'+str(pk))


class CatalogoListView(BasicEMixin, ListView):

    template_name = 'maestro/catalogo-list.html'
    model = Producto
    nav_name = 'nav_catalogo'
    nav_main = 'nav_main_producto'
    view_name = 'catalogo'
    action_name = 'leer'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['catalogo_filtro'] = CatalogoFiltroForm(self.request.GET)
        return context

    def get_queryset(self):
        sucursales = self.request.GET.getlist('sucursal')
        if len(sucursales) > 0:
            query = Producto.objects.filter(catalogo__in=sucursales)
        else:
            query = Producto.objects.none()
        return query


class CatalogoDeleteView(BasicEMixin, RedirectView):

    url = '/maestro/catalogo/'
    view_name = 'catalogo'
    action_name = 'eliminar'

    def get_redirect_url(self, *args, **kwargs):
        sucursal = Sucursal.objects.get(pk=self.request.POST['sucursal'])
        producto = Producto.objects.get(pk=self.request.POST['producto'])
        producto.catalogo.remove(sucursal)
        url = self.url + self.request.POST['sucursal']
        return url


class CatalogoAddView(BasicEMixin, TemplateView):

    template_name = 'maestro/catalogo-add.html'
    nav_name = 'nav_catalogo'
    nav_main = 'nav_main_producto'
    view_name = 'catalogo'
    action_name = 'crear'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sucursales'] = Sucursal.objects.all()
        context['own_sucursal'] = int(self.kwargs['pk'])
        return context

    def post(self, request, *args, **kwargs):
        sucursal = Sucursal.objects.get(pk=self.request.POST['sucursal'])
        if request.POST['catalogo_to_save'] != "":
            catalogo_toadd = request.POST['catalogo_to_save'].split(',')
            for c in catalogo_toadd:
                producto = Producto.objects.get(pk=self.request.POST['p'+c+'-presentacion'])
                producto.catalogo.add(sucursal)
        return redirect('/maestro/catalogo/?sucursal='+str(sucursal.id))


class ProveedorListView(BasicEMixin, ListView):

    template_name = 'maestro/proveedor-list.html'
    model = Proveedor
    nav_name = 'nav_proveedor'
    nav_main = 'nav_main_proveedor'
    view_name = 'proveedor'
    action_name = 'leer'

    def get_queryset(self):
        query = super().get_queryset()
        query = query.filter(empresa__in=empresa_list(self.request.user))
        return query


class ProveedorDetailView(BasicEMixin, DetailView):

    template_name = 'maestro/proveedor-detail.html'
    model = Proveedor
    nav_name = 'nav_proveedor'
    nav_main = 'nav_main_proveedor'
    view_name = 'proveedor'
    action_name = 'leer'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['catalogo'] = CatalogoxProveedor.objects.filter(proveedor=self.kwargs['pk'])
        return context


class ProveedorEditView(BasicEMixin, TemplateView):

    template_name = 'maestro/proveedor-edit.html'
    nav_name = 'nav_proveedor'
    nav_main = 'nav_main_proveedor'
    view_name = 'proveedor'
    action_name = 'editar'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.kwargs['pk'] == 0:
            context['object'] = ProveedorForm(user=self.request.user)
        else:
            proveedor = Proveedor.objects.get(pk=self.kwargs['pk'])
            context['object'] = ProveedorForm(instance=proveedor, user=self.request.user)
        return context

    def post(self, request, *args, **kwargs):
        if self.kwargs['pk'] == 0:
            form = ProveedorForm(request.POST, user=self.request.user)
        else:
            proveedor = Proveedor.objects.get(pk=self.kwargs['pk'])
            form = ProveedorForm(request.POST, instance=proveedor, user=self.request.user)
        if form.is_valid():
            form.save()
        else:
            return HttpResponse(form.errors)
        return redirect('/maestro/proveedor')


class CatalogoProveedorListView(BasicEMixin, ListView):

    template_name = 'maestro/catalogoproveedor-list.html'
    model = Producto
    nav_name = 'nav_catalogoproveedor'
    nav_main = 'nav_main_proveedor'
    view_name = 'catalogo_proveedor'
    action_name = 'leer'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['catalogo_filtro'] = CatalogoProveedorFiltroForm(self.request.GET, user=self.request.user)
        return context

    def get_queryset(self):
        proveedor = self.request.GET.getlist('proveedor')
        if len(proveedor) > 0:
            query = CatalogoxProveedor.objects.filter(proveedor__in=proveedor,
                                                      proveedor__empresa__in=empresa_list(self.request.user))
        else:
            query = CatalogoxProveedor.objects.none()
        return query


class CatalogoProveedorDeleteView(BasicEMixin, RedirectView):

    url = '/maestro/catalogoproveedor/'
    view_name = 'catalogo_proveedor'
    action_name = 'eliminar'

    def get_redirect_url(self, *args, **kwargs):
        proveedor = self.request.POST['proveedor']
        producto = self.request.POST['producto']
        CatalogoxProveedor.objects.get(proveedor=proveedor, producto=producto).delete()
        url = self.url + self.request.POST['proveedor']
        return url


class CatalogoProveedorAddView(BasicEMixin, TemplateView):

    template_name = 'maestro/catalogoproveedor-add.html'
    nav_name = 'nav_catalogoproveedor'
    nav_main = 'nav_main_proveedor'
    view_name = 'catalogo_proveedor'
    action_name = 'crear'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['proveedores'] = Proveedor.objects.all()
        context['own_proveedor'] = int(self.kwargs['pk'])
        return context

    def post(self, request, *args, **kwargs):
        proveedor = Proveedor.objects.get(pk=self.request.POST['proveedor'])
        if request.POST['catalogo_to_save'] != "":
            catalogo_toadd = request.POST['catalogo_to_save'].split(',')
            for c in catalogo_toadd:
                try:
                    CatalogoxProveedor.objects.get(proveedor=self.request.POST['proveedor'],
                                                   producto=self.request.POST['p'+c+'-producto'])
                except CatalogoxProveedor.DoesNotExist:
                    form = CatalogoProveedorForm(request.POST, prefix='p'+c)
                    if form.is_valid():
                        catalogoproveedor = form.save(commit=False)
                        catalogoproveedor.proveedor = proveedor
                        catalogoproveedor.save()
        return redirect('/maestro/catalogoproveedor/?proveedor='+str(proveedor.id))


class UsuarioListView(BasicEMixin, ListView):

    template_name = 'maestro/usuario-list.html'
    model = User
    nav_name = 'nav_usuario'
    nav_main = 'nav_main_usuario'
    view_name = 'usuario'
    action_name = 'leer'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

    def get_queryset(self):
        query = User.objects.all()
        return query


class UsuarioDetailView(BasicEMixin, DetailView):
    template_name = 'maestro/usuario-detail.html'
    model = User
    nav_name = 'nav_usuario'
    nav_main = 'nav_main_usuario'
    view_name = 'usuario'
    action_name = 'leer'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        return context


class UsuarioEditView(BasicEMixin, TemplateView):

    template_name = 'maestro/usuario-edit.html'
    nav_name = 'nav_usuario'
    nav_main = 'nav_main_usuario'
    view_name = 'usuario'
    action_name = 'editar'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.kwargs['pk'] == 0:
            context['object'] = UsuarioForm
        else:
            user = User.objects.get(pk=self.kwargs['pk'])
            grupoassign = AsignacionGrupo.objects.get(usuario=user)
            context['object'] = UsuarioForm(instance=user, initial={'grupo': grupoassign.grupo,
                                                                    'sucursal': grupoassign.sucursal.all()})
        return context

    def post(self, request, *args, **kwargs):
        if self.kwargs['pk'] == 0:
            form = UsuarioForm(request.POST)
            if form.is_valid():
                user = form.save(commit=False)
                user.password = make_password(form.cleaned_data['password'])
                user.save()
                try:
                    grupoassign = AsignacionGrupo.objects.get(usuario=user)
                    grupoassign.grupo = form.cleaned_data['grupo']
                    grupoassign.sucursal = form.cleaned_data['sucursal']
                    grupoassign.save()
                except AsignacionGrupo.DoesNotExist:
                    AsignacionGrupo(usuario=user, grupo=form.cleaned_data['grupo']).save()
            else:
                return HttpResponse(form.errors)
        else:
            user = User.objects.get(pk=self.kwargs['pk'])
            form = UsuarioForm(request.POST, instance=user)
            if form.is_valid():
                user = form.save(commit=False)
                user.password = make_password(form.cleaned_data['password'])
                user.save()
                try:
                    grupoassign = AsignacionGrupo.objects.get(usuario=user)
                    grupoassign.grupo = form.cleaned_data['grupo']
                    grupoassign.sucursal.set(form.cleaned_data['sucursal'])
                    grupoassign.save()
                except AsignacionGrupo.DoesNotExist:
                    AsignacionGrupo(usuario=user, grupo=form.cleaned_data['grupo']).save()
            else:
                return HttpResponse(form.errors)
        return redirect('/maestro/usuario/'+str(user.id))

# from ventas.models import OfertaVenta
# from almacen.models import Stock
# from maestro.models import Vista, AccionxVista, AsignacionAccion, Accion, Grupo
#
#
# def migracion(request):
#     # libro = Workbook()
#     # libro = load_workbook("./migracion.xlsx")
#     # h = libro["Hoja1"]
#     # direccion = [['K', 'L'], ['M', 'N'], ['O', 'P'], ['Q', 'R']]
#     # empresa = Empresa.objects.get(pk=1)
#     # sucursal = Sucursal.objects.get(pk=1)
#     # almacen = Almacen.objects.get(pk=1)
#     # for index in range(5, 337):
#     #     presentaciones = []
#     #     # Guardar producto
#     #     name_producto = h['J'+str(index)].value
#     #     try:
#     #         producto = Producto.objects.get(descripcion=name_producto)
#     #     except Producto.DoesNotExist:
#     #         producto = Producto(descripcion=name_producto, empresa=empresa)
#     #         producto.save()
#     #     producto.catalogo.add(sucursal)
#     #     for i in range(0, 4):
#     #         direccion_name = direccion[i][0]
#     #         direccion_detail = direccion[i][1]
#     #         name = h[direccion_name+str(index)].value
#     #         detail = h[direccion_detail+str(index)].value
#     #         if (name != '' and name is not None) and (detail != '' and detail is not None):
#     #             detail_split = detail.split(' ')
#     #             if detail_split[0] != '' and detail_split[1] != '':
#     #                 if detail_split[1][-1:] == 'S':
#     #                     detail_split[1] = detail_split[1][:-1]
#     #                 if name[-1:] == 'S':
#     #                     name = name[:-1]
#     #                 presentaciones.append([name, detail_split[0], detail_split[1]])
#     #     for idx, p in enumerate(reversed(presentaciones)):
#     #         if idx == 0:
#     #             if p[2] == 'U':
#     #                 p[2] = 'UNIDAD'
#     #             try:
#     #                 presentacion_temp = Presentacion.objects.get(descripcion=p[2])
#     #             except Presentacion.DoesNotExist:
#     #                 presentacion_temp = Presentacion(descripcion=p[2])
#     #                 presentacion_temp.save()
#     #             PresentacionxProducto(presentacion=presentacion_temp, producto=producto, cantidad=1).save()
#     #         try:
#     #             presentacion_to_save = Presentacion.objects.get(descripcion=p[0])
#     #         except Presentacion.DoesNotExist:
#     #             presentacion_to_save = Presentacion(descripcion=p[0])
#     #             presentacion_to_save.save()
#     #         presentacion_prev = PresentacionxProducto.objects.get(producto=producto, presentacion__descripcion=p[2])
#     #         try:
#     #             PresentacionxProducto.objects.get(presentacion=presentacion_to_save, producto=producto)
#     #         except PresentacionxProducto.DoesNotExist:
#     #             PresentacionxProducto(presentacion=presentacion_to_save, producto=producto,
#     #                                   cantidad=int(p[1])*presentacion_prev.cantidad).save()
#     #     # # Calcular el precio
#     #     #
#     #     #     # Precio Venta Alto
#     #     # name_presentacion_precio = h['W'+str(index)].value
#     #     # precio = h['V'+str(index)].value
#     #     # if name_presentacion_precio != '' and name_presentacion_precio is not None:
#     #     #     try:
#     #     #         presentacion_precio = PresentacionxProducto.objects.get(producto=producto,
#     #     #                                                                 presentacion__descripcion=name_presentacion_precio)
#     #     #         producto.precio_venta = round((precio / presentacion_precio.cantidad), 1)
#     #     #         if (producto.precio_venta * presentacion_precio.cantidad) > precio:
#     #     #             OfertaVenta(sucursal=sucursal, tipo='2', tipo_duracion='2', producto_oferta=producto,
#     #     #                         presentacion_oferta=presentacion_precio, cantidad_oferta=1,
#     #     #                         cantidad_unidad_oferta=presentacion_precio.cantidad,
#     #     #                         retorno=(producto.precio_venta * presentacion_precio.cantidad)-precio).save()
#     #     #     except PresentacionxProducto.DoesNotExist:
#     #     #         pass
#     #     #
#     #     #     # Precio Compra
#     #     # precio_compra = h['T'+str(index)].value
#     #     # if precio_compra != '' and precio_compra is not None:
#     #     #     presentacion_precio_compra = PresentacionxProducto.objects.get(producto=producto,
#     #     #                                                                    presentacion__descripcion=presentaciones[0][0])
#     #     #     producto.precio_compra = round((precio_compra / presentacion_precio_compra.cantidad), 2)
#     #     #
#     #     #     # Precio Venta Bajo
#     #     # precio_venta = h['U'+str(index)].value
#     #     # if precio_venta != '' and precio_venta is not None:
#     #     #     presentacion_precio_venta = PresentacionxProducto.objects.get(producto=producto,
#     #     #                                                                   presentacion__descripcion=presentaciones[0][0])
#     #     #     if (producto.precio_venta * presentacion_precio_venta.cantidad) > precio_venta:
#     #     #         ofertaventa = OfertaVenta.objects.filter(producto_oferta=producto, is_active=True, estado=True)
#     #     #         descuento = 0
#     #     #         for ofv in ofertaventa:
#     #     #             descuento += (presentacion_precio_venta.cantidad/ofv.presentacion_oferta.cantidad) * float(ofv.retorno)
#     #     #         OfertaVenta(sucursal=sucursal, tipo='2', tipo_duracion='2', producto_oferta=producto,
#     #     #                     presentacion_oferta=presentacion_precio_venta, cantidad_oferta=1,
#     #     #                     cantidad_unidad_oferta=presentacion_precio_venta.cantidad,
#     #     #                     retorno=((producto.precio_venta * presentacion_precio_venta.cantidad)-precio_venta)-descuento)\
#     #     #             .save()
#     #     # if producto.precio_venta != 0 and producto.precio_compra != 0:
#     #     #     producto.utilidad_monetaria = producto.precio_venta-producto.precio_compra
#     #     # producto.save()
#     #
#     #     # Stock
#     #     direccion_stock = [['B', 'C'], ['D', 'E'], ['F', 'G'], ['H', 'I']]
#     #     try:
#     #         stock = Stock.objects.get(almacen=almacen, producto=producto)
#     #     except Stock.DoesNotExist:
#     #         stock = Stock(almacen=almacen, producto=producto, cantidad=0)
#     #         stock.save()
#     #     for j in range(0, 4):
#     #         direccion_stock_name = direccion_stock[j][1]
#     #         direccion_stock_detail = direccion_stock[j][0]
#     #         name_stock = h[direccion_stock_name+str(index)].value
#     #         cantidad_stock = h[direccion_stock_detail+str(index)].value
#     #         if (name_stock != '' and name_stock is not None) and (cantidad_stock != '' and cantidad_stock is not None):
#     #             if name_stock[-1:] == 'S':
#     #                 name_stock = name_stock[:-1]
#     #             presentacion_stock = PresentacionxProducto.objects.get(producto=producto,
#     #                                                                    presentacion__descripcion=name_stock)
#     #             stock.cantidad += presentacion_stock.cantidad * int(cantidad_stock)
#     #     stock.save()
#     # grupo = Grupo.objects.get(pk=1)
#     # vista = Vista.objects.get(pk=23)
#     # accion = Accion.objects.filter(id__in=[1, 2, 4, 9, 10])
#     # for a in accion:
#     #     try:
#     #         axv = AccionxVista.objects.get(accion=a.id, vista=vista.id)
#     #     except AccionxVista.DoesNotExist:
#     #         axv = AccionxVista(accion=a, vista=vista)
#     #         axv.save()
#     #     try:
#     #         AsignacionAccion.objects.get(grupo=grupo.id, accionxvista=axv.id)
#     #     except AsignacionAccion.DoesNotExist:
#     #         aa = AsignacionAccion(grupo=grupo, accionxvista=axv)
#     #         aa.save()
#     productos = Producto.objects.all()
#     proveedor = Proveedor.objects.get(pk=1)
#     for p in productos:
#         CatalogoxProveedor(producto=p, proveedor=proveedor).save()
#     return HttpResponse('Done')
